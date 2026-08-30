from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .column_mapping import detect_column_mapping, normalize_header
from .config import CollectorConfig
from .errors import ValidationError


SUPPORTED_SUFFIXES = {".csv", ".xlsx"}


def discover_input_files(config: CollectorConfig) -> list[Path]:
    input_dir = config.path("input_dir")
    if not input_dir.exists():
        raise ValidationError(f"No existe input_dir: {input_dir}")
    files = [
        path
        for path in sorted(input_dir.iterdir())
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES
    ]
    return files


def read_file(path: Path, config: CollectorConfig) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path, config)
    if path.suffix.lower() == ".xlsx":
        return read_xlsx(path, config)
    raise ValidationError(f"Tipo de archivo no soportado: {path.name}")


def read_csv(path: Path, config: CollectorConfig) -> list[dict[str, Any]]:
    configured_encoding = str(config.raw.get("encoding", "utf-8"))
    encoding = "utf-8-sig" if configured_encoding.lower().replace("_", "-") == "utf-8" else configured_encoding
    try:
        with path.open("r", encoding=encoding, newline="") as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;").delimiter
            except csv.Error:
                delimiter = str(config.raw.get("delimiter", ","))
            reader = csv.DictReader(fh, delimiter=delimiter)
            fieldnames = [str(value or "").lstrip("\ufeff").strip() for value in (reader.fieldnames or [])]
            if not fieldnames or (len(fieldnames) == 1 and any(mark in fieldnames[0] for mark in (",", ";"))):
                raise ValidationError("No pude reconocer las columnas del CSV ni su separador.")
            rows = []
            for raw in reader:
                row = {fieldnames[index]: value for index, value in enumerate(raw.values()) if index < len(fieldnames) and fieldnames[index]}
                rows.append(row)
            return rows
    except (OSError, UnicodeError) as exc:
        raise ValidationError(f"No pude leer {path.name}; puede estar abierto o usar una codificacion no soportada.") from exc


def _xlsx_header_score(row: tuple[Any, ...], config: CollectorConfig | None) -> tuple[int, float, int]:
    headers = [str(value).lstrip("\ufeff").strip() for value in row if value not in (None, "")]
    if len(headers) < 2:
        return (0, 0.0, len(headers))
    detected = detect_column_mapping(headers)
    configured_sources = {
        normalize_header(source)
        for source in ((config.column_mapping if config else {}) or {}).values()
        if str(source).strip()
    }
    configured_matches = sum(normalize_header(header) in configured_sources for header in headers)
    semantic_matches = len(detected.mapping) + configured_matches
    return (semantic_matches, detected.confidence, len(headers))


def read_xlsx(path: Path, config: CollectorConfig | None = None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl no está disponible para leer XLSX") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ValidationError(f"No pude leer {path.name}; puede estar abierto o no ser un XLSX valido.") from exc
    try:
        configured_sheet = str((config.raw if config else {}).get("xlsx_sheet", "")).strip()
        sheets = [workbook[configured_sheet]] if configured_sheet and configured_sheet in workbook.sheetnames else list(workbook.worksheets)
        candidates: list[tuple[tuple[int, float, int], int, Any, list[tuple[Any, ...]]]] = []
        for sheet in sheets:
            rows = list(sheet.iter_rows(values_only=True))
            scored = [(_xlsx_header_score(row, config), index) for index, row in enumerate(rows[:25])]
            scored = [(score, index) for score, index in scored if score[2] >= 2]
            if not scored:
                continue
            score, header_index = max(scored, key=lambda item: (item[0], -item[1]))
            candidates.append((score, -header_index, sheet, rows))
        if candidates:
            _score, negative_header_index, _sheet, rows = max(candidates, key=lambda item: (item[0], item[1]))
            header_index = -negative_header_index
            headers = [str(value).lstrip("\ufeff").strip() if value is not None else "" for value in rows[header_index]]
            result: list[dict[str, Any]] = []
            for row in rows[header_index + 1 :]:
                item = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
                if any(value not in (None, "") for value in item.values()):
                    result.append(item)
            return result
        return []
    finally:
        workbook.close()


def detect_columns(files: list[Path], config: CollectorConfig) -> set[str]:
    columns: set[str] = set()
    for path in files:
        rows = read_file(path, config)
        if rows:
            columns.update(rows[0].keys())
        elif path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                sample = fh.read(8192)
                fh.seek(0)
                try:
                    delimiter = csv.Sniffer().sniff(sample, delimiters=",;").delimiter
                except csv.Error:
                    delimiter = str(config.raw.get("delimiter", ","))
                reader = csv.DictReader(fh, delimiter=delimiter)
                columns.update(str(value or "").lstrip("\ufeff").strip() for value in (reader.fieldnames or []))
    return columns
