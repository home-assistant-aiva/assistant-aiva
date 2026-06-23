from __future__ import annotations

import argparse
import csv
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from aiva_collector.column_mapping import detect_column_mapping


ROOT = Path(__file__).resolve().parents[1]
CASES_DIR = ROOT / "samples" / "mapping_cases"
REPORT_PATH = CASES_DIR / "output" / "mapping_validation_report.json"

HIGH_CASES = {
    "technical_columns.xlsx",
    "simple_columns.xlsx",
    "pos_spanish_columns.xlsx",
    "accented_columns.csv",
}
MIN_CONFIDENCE = {
    "english_columns.xlsx": 0.80,
    "messy_but_mappable.xlsx": 0.75,
}
LOW_CASE = "low_confidence_columns.xlsx"
MISSING_REQUIRED_CASE = "missing_required_columns.xlsx"


@dataclass
class CaseReport:
    file: str
    status: str
    confidence: float
    missing_required: list[str]
    mapped_fields: dict[str, str]
    warnings: list[str]


def read_headers(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh).fieldnames or [])
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            row = next(workbook.active.iter_rows(values_only=True), ())
            return [str(value).strip() for value in row if value is not None and str(value).strip()]
        finally:
            workbook.close()
    return []


def validate_case(path: Path) -> CaseReport:
    result = detect_column_mapping(read_headers(path))
    return CaseReport(
        file=path.name,
        status=result.status,
        confidence=result.confidence,
        missing_required=result.missing_required,
        mapped_fields=result.mapping,
        warnings=result.warnings,
    )


def validate_reports(reports: list[CaseReport]) -> list[str]:
    failures: list[str] = []
    by_file = {report.file: report for report in reports}
    for name in HIGH_CASES:
        report = by_file[name]
        if report.status != "auto_approved" or report.confidence < 0.85 or report.missing_required:
            failures.append(f"{name}: expected high confidence auto_approved, got {report.status} {report.confidence}")
    for name, min_confidence in MIN_CONFIDENCE.items():
        report = by_file[name]
        if report.confidence < min_confidence or report.missing_required:
            failures.append(f"{name}: expected confidence >= {min_confidence}, got {report.confidence}")
    low = by_file[LOW_CASE]
    if low.status == "auto_approved":
        failures.append(f"{LOW_CASE}: must not auto approve")
    missing = by_file[MISSING_REQUIRED_CASE]
    if missing.status != "failed":
        failures.append(f"{MISSING_REQUIRED_CASE}: expected failed, got {missing.status}")
    for required in ("cantidad_vendida", "precio_venta"):
        if required not in missing.missing_required:
            failures.append(f"{MISSING_REQUIRED_CASE}: missing_required does not include {required}")
    for report in reports:
        sources = list(report.mapped_fields.values())
        if len(sources) != len(set(sources)):
            failures.append(f"{report.file}: same source column mapped more than once")
    return failures


def print_table(reports: list[CaseReport]) -> None:
    print(f"{'file':34} {'status':14} {'confidence':10} {'missing_required':34} mapped fields")
    print("-" * 118)
    for report in reports:
        mapped = ", ".join(f"{field}->{source}" for field, source in sorted(report.mapped_fields.items()))
        missing = ",".join(report.missing_required) or "-"
        print(f"{report.file:34} {report.status:14} {report.confidence:<10.3f} {missing:34} {mapped}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Validate synthetic AIVA column mapping cases without backend calls")
    parser.add_argument("--cases-dir", default=str(CASES_DIR))
    parser.add_argument("--report", default=str(REPORT_PATH))
    args = parser.parse_args(argv)

    cases_dir = Path(args.cases_dir)
    paths = sorted(path for path in cases_dir.iterdir() if path.suffix.lower() in {".csv", ".xlsx"})
    reports = [validate_case(path) for path in paths]
    print_table(reports)

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    payload: dict[str, Any] = {"ok": True, "cases": [asdict(report) for report in reports], "failures": []}
    failures = validate_reports(reports)
    payload["ok"] = not failures
    payload["failures"] = failures
    report_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    if failures:
        print("\nFAILURES:")
        for failure in failures:
            print(f"- {failure}")
        return 1
    print(f"\nMapping validation OK: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
