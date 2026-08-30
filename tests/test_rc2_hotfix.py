from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook

from aiva_collector.cli import _filter_unchanged_read_only_files, _single_run_lock, main
from aiva_collector.config import CollectorConfig, load_config
from aiva_collector.desktop_service import export_diagnostics
from aiva_collector.local_state import connect, local_db_path, update_file_state, upsert_detected_file
from aiva_collector.readers import read_file


EXPECTED_MAPPING = {
    "producto_nombre": "Producto",
    "fecha": "Fecha",
    "cantidad_vendida": "Cantidad",
    "stock_actual": "Stock",
    "precio_venta": "Precio",
    "costo_unitario": "Costo",
}


def _config(tmp_path: Path, *, commerce_id: str = "commerce-one", collector_id: str = "collector-one") -> Path:
    data = json.loads(Path("configs/example_config.json").read_text(encoding="utf-8"))
    data.update(
        {
            "collector_version": "0.2.7rc2",
            "backend_url": "http://backend.test:8080",
            "commerce_id": commerce_id,
            "collector_id": collector_id,
            "input_dir": str(tmp_path / "entrada"),
            "processed_dir": str(tmp_path / "procesados"),
            "error_dir": str(tmp_path / "rechazados"),
            "output_dir": str(tmp_path / "output"),
            "state_dir": str(tmp_path / "estado"),
            "log_file": str(tmp_path / "logs" / "collector.log"),
            "source_read_only": True,
            "move_processed_files": False,
            "move_error_files": False,
            "keep_original_files": True,
            "stable_file_interval_seconds": 0,
        }
    )
    (tmp_path / "entrada").mkdir(exist_ok=True)
    path = tmp_path / "config.windows.json"
    path.write_text(json.dumps(data), encoding="utf-8")
    return path


def _write_exact_csv(path: Path, *, delimiter: str = ",", bom: bool = False) -> None:
    text = delimiter.join(("Producto", "Fecha", "Cantidad", "Stock", "Precio", "Costo")) + "\n"
    text += delimiter.join(("Yerba", "2026-08-30", "2", "8", "2500", "1800")) + "\n"
    path.write_text(("\ufeff" if bom else "") + text, encoding="utf-8")


def _mock_backend(monkeypatch, sent: list[dict], candidates: list[dict]) -> None:
    monkeypatch.setenv("AIVA_COLLECTOR_TOKEN", "synthetic-token")
    monkeypatch.setattr("aiva_collector.cli._backend_mapping", lambda config: None)
    monkeypatch.setattr("aiva_collector.cli.CollectorClient.post_status", lambda self, status, message=None: {})

    def post_candidate(self, payload):
        candidates.append(payload)
        return {"candidate": {"id": "candidate-test"}}

    def send_summary(self, payload):
        sent.append(payload)
        return {"summary_id": "summary-test", "_http_status_code": 201}

    monkeypatch.setattr("aiva_collector.cli.CollectorClient.post_mapping_candidate", post_candidate)
    monkeypatch.setattr("aiva_collector.cli.CollectorClient.send_summary", send_summary)


@pytest.mark.parametrize(("delimiter", "bom"), [(",", False), (",", True), (";", False), (";", True)])
def test_csv_variants_detect_exact_mapping(tmp_path, monkeypatch, delimiter, bom):
    config_path = _config(tmp_path)
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source, delimiter=delimiter, bom=bom)
    sent: list[dict] = []
    candidates: list[dict] = []
    _mock_backend(monkeypatch, sent, candidates)

    assert main(["run-auto", "--config", str(config_path)]) == 0

    assert len(sent) == 1
    assert candidates[0]["status"] == "auto_approved"
    assert candidates[0]["suggested_mapping"] == EXPECTED_MAPPING


def test_xlsx_displaced_header_detects_exact_mapping(tmp_path):
    config_path = _config(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append(["Reporte de prueba"])
    sheet.append(list(EXPECTED_MAPPING.values()))
    sheet.append(["Yerba", "2026-08-30", 2, 8, 2500, 1800])
    source = tmp_path / "entrada" / "ventas.xlsx"
    workbook.save(source)

    rows = read_file(source, load_config(config_path))

    assert list(rows[0]) == list(EXPECTED_MAPPING.values())


def test_xlsx_header_selection_prefers_semantic_compatibility_across_sheets(tmp_path):
    config_path = _config(tmp_path)
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Portada"
    cover.append(["Comercio Demo", "Reporte mensual"])
    data = workbook.create_sheet("Ventas")
    data.append(["Periodo", "Agosto"])
    data.append(list(EXPECTED_MAPPING.values()))
    data.append(["Yerba", "2026-08-30", 2, 8, 2500, 1800])
    source = tmp_path / "entrada" / "ventas-varias-hojas.xlsx"
    workbook.save(source)

    rows = read_file(source, load_config(config_path))

    assert list(rows[0]) == list(EXPECTED_MAPPING.values())
    assert rows[0]["Producto"] == "Yerba"


def test_full_sync_sends_once_then_explains_duplicate(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    _write_exact_csv(tmp_path / "entrada" / "ventas.csv")
    sent: list[dict] = []
    candidates: list[dict] = []
    _mock_backend(monkeypatch, sent, candidates)

    assert main(["run-auto", "--config", str(config_path)]) == 0
    first = json.loads((tmp_path / "estado" / "last_auto_run.json").read_text(encoding="utf-8"))
    assert first["files_found"] == first["files_eligible"] == first["files_processed"] == 1
    assert first["summaries_sent"] == 1

    assert main(["run-auto", "--config", str(config_path)]) == 0
    second = json.loads((tmp_path / "estado" / "last_auto_run.json").read_text(encoding="utf-8"))
    assert len(sent) == 1
    assert second["files_found"] == 1
    assert second["files_eligible"] == 0
    assert second["files_skipped"] == 1
    assert second["skipped_details"][0]["reason"] == "El archivo ya fue enviado a este comercio."


def test_ambiguous_mapping_is_reported_then_saved_mapping_reprocesses_same_file(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    source = tmp_path / "entrada" / "ambiguo.csv"
    source.write_text(
        "Descripción del producto,Cantidad venta,Precio ventas\nYerba,2,2500\n",
        encoding="utf-8",
    )
    sent: list[dict] = []
    candidates: list[dict] = []
    _mock_backend(monkeypatch, sent, candidates)

    assert main(["run-auto", "--config", str(config_path)]) == 0
    assert sent == []
    assert candidates[0]["status"] == "needs_review"
    state = json.loads((tmp_path / "estado" / "last_auto_run.json").read_text(encoding="utf-8"))
    assert state["needs_review"] == 1

    approved = candidates[0]["suggested_mapping"]
    monkeypatch.setattr("aiva_collector.cli._backend_mapping", lambda config: approved)
    assert main(["run-auto", "--config", str(config_path)]) == 0
    assert len(sent) == 1


def test_same_sha_sent_for_other_activation_is_eligible(tmp_path):
    config_path = _config(tmp_path, commerce_id="commerce-new", collector_id="collector-new")
    config = load_config(config_path)
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    conn = connect(local_db_path(config))
    try:
        upsert_detected_file(
            conn,
            file_id="old-activation",
            commerce_id="commerce-old",
            collector_id="collector-old",
            backend_url="http://backend.test:8080",
            path=source,
            file_sha256=digest,
        )
        update_file_state(conn, "old-activation", status="sent")

        assert _filter_unchanged_read_only_files(config, conn, [source]) == [source]
    finally:
        conn.close()


def test_abandoned_processing_is_recovered(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    config = load_config(config_path)
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    conn = connect(local_db_path(config))
    try:
        upsert_detected_file(
            conn,
            file_id="interrupted",
            commerce_id=config.commerce_id,
            collector_id=config.collector_id,
            backend_url=config.backend_url,
            path=source,
            file_sha256=digest,
            status="processing",
        )
        update_file_state(
            conn,
            "interrupted",
            lease_expires_at=(datetime.now(timezone.utc) - timedelta(minutes=1)).isoformat(),
        )
    finally:
        conn.close()
    sent: list[dict] = []
    _mock_backend(monkeypatch, sent, [])

    assert main(["run-auto", "--config", str(config_path)]) == 0
    assert len(sent) == 1
    conn = connect(local_db_path(config))
    try:
        assert conn.execute("SELECT status FROM processed_files WHERE file_id = 'interrupted'").fetchone()[0] == "sent"
        assert conn.execute("SELECT COUNT(*) FROM processed_file_events WHERE event_type = 'abandoned_processing_recovered'").fetchone()[0] == 1
    finally:
        conn.close()


def test_active_processing_lease_is_skipped_with_reason(tmp_path):
    config_path = _config(tmp_path)
    config = load_config(config_path)
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source)
    conn = connect(local_db_path(config))
    try:
        upsert_detected_file(
            conn,
            file_id="active",
            commerce_id=config.commerce_id,
            collector_id=config.collector_id,
            backend_url=config.backend_url,
            path=source,
            file_sha256="sha",
            status="processing",
        )
        update_file_state(conn, "active", lease_expires_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat())
        reasons: list[dict[str, str]] = []
        assert _filter_unchanged_read_only_files(config, conn, [source], skipped_details=reasons) == []
        assert "procesamiento activo" in reasons[0]["reason"]
    finally:
        conn.close()


def test_active_run_lock_records_real_skip_reason(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source)
    lock = tmp_path / "estado" / "aiva_collector.lock"
    lock.parent.mkdir()
    lock.write_text(json.dumps({"pid": 123, "started_at": datetime.now(timezone.utc).isoformat()}), encoding="utf-8")
    monkeypatch.setenv("AIVA_COLLECTOR_TOKEN", "synthetic-token")

    assert main(["run-auto", "--config", str(config_path)]) == 0

    state = json.loads((tmp_path / "estado" / "last_auto_run.json").read_text(encoding="utf-8"))
    assert state["files_found"] == state["files_skipped"] == 1
    assert state["files_processed"] == 0
    assert state["skipped_details"][0]["status"] == "lock_active"
    assert "sincronizacion activa" in state["error_summary"]


def test_header_only_file_is_rejected_with_explanation(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    source = tmp_path / "entrada" / "sin-registros.csv"
    source.write_text("Producto,Fecha,Cantidad,Stock,Precio,Costo\n", encoding="utf-8")
    sent: list[dict] = []
    _mock_backend(monkeypatch, sent, [])

    assert main(["run-auto", "--config", str(config_path)]) == 2

    state = json.loads((tmp_path / "estado" / "last_auto_run.json").read_text(encoding="utf-8"))
    assert sent == []
    assert state["rejected"] == 1
    assert "no contiene registros" in state["error_summary"]


def test_stale_lock_is_recovered(tmp_path):
    config = CollectorConfig(raw={"state_dir": str(tmp_path / "estado")}, config_path=tmp_path / "config.json")
    lock = tmp_path / "estado" / "aiva_collector.lock"
    lock.parent.mkdir()
    lock.write_text(json.dumps({"pid": 1, "started_at": (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()}), encoding="utf-8")

    with _single_run_lock(config) as acquired:
        assert acquired is True

    assert not lock.exists()


def test_diagnostic_zip_excludes_source_content_and_redacts_secret(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    secret = "secret-value-not-for-zip"
    config_data = json.loads(config_path.read_text(encoding="utf-8"))
    config_data["collector_token"] = secret
    config_path.write_text(json.dumps(config_data), encoding="utf-8")
    source = tmp_path / "entrada" / "ventas.csv"
    _write_exact_csv(source)
    (tmp_path / "logs").mkdir()
    commercial_secret = "YERBA-COMERCIAL-NO-DEBE-SALIR"
    (tmp_path / "logs" / "collector.log").write_text(
        f"Authorization: Bearer {secret}\n"
        f'payload={{"Producto":"{commercial_secret}","Cantidad":2,"Stock":8,"Precio":2500,"Costo":1800}}\n',
        encoding="utf-8",
    )
    monkeypatch.setenv("AIVA_COLLECTOR_STANDARD_CONFIG", str(config_path))
    monkeypatch.setenv("AIVA_COLLECTOR_DATA_DIR", str(tmp_path))
    connect(tmp_path / "estado" / "aiva_collector.db").close()

    result = export_diagnostics()

    assert result.ok is True
    zip_path = tmp_path / "diagnostico" / "aiva-collector-diagnostico-rc2.zip"
    with zipfile.ZipFile(zip_path) as archive:
        assert "ventas.csv" not in archive.namelist()
        combined = b"".join(archive.read(name) for name in archive.namelist())
    assert secret.encode() not in combined
    assert commercial_secret.encode() not in combined
    assert hashlib.sha256(source.read_bytes()).hexdigest().encode() in combined


def test_diagnostic_export_reads_pre_rc2_database_without_migrating_it(tmp_path, monkeypatch):
    config_path = _config(tmp_path)
    db_path = tmp_path / "estado" / "aiva_collector.db"
    db_path.parent.mkdir()
    with sqlite3.connect(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE processed_files (
                file_id TEXT PRIMARY KEY, file_path TEXT NOT NULL, file_name TEXT NOT NULL,
                file_size INTEGER, file_mtime TEXT, file_sha256 TEXT NOT NULL,
                detected_at TEXT NOT NULL, status TEXT NOT NULL, created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE UNIQUE INDEX idx_processed_files_sha256 ON processed_files(file_sha256);
            INSERT INTO processed_files VALUES (
                'legacy', 'C:/private/ventas.csv', 'ventas.csv', 123, '2026-08-30',
                'legacy-sha', '2026-08-30', 'sent', '2026-08-30', '2026-08-30'
            );
            """
        )
    monkeypatch.setenv("AIVA_COLLECTOR_STANDARD_CONFIG", str(config_path))
    monkeypatch.setenv("AIVA_COLLECTOR_DATA_DIR", str(tmp_path))

    result = export_diagnostics()

    assert result.ok is True
    with zipfile.ZipFile(tmp_path / "diagnostico" / "aiva-collector-diagnostico-rc2.zip") as archive:
        diagnostic = json.loads(archive.read("diagnostic.json"))
    assert diagnostic["local_state"]["quick_check"] == "ok"
    assert diagnostic["local_state"]["processed_files"][0]["file_id"] == "legacy"
    with sqlite3.connect(db_path) as conn:
        assert "backend_url" not in {row[1] for row in conn.execute("PRAGMA table_info(processed_files)")}


def test_local_state_migrates_global_sha_index_without_losing_rows(tmp_path):
    db_path = tmp_path / "estado" / "aiva_collector.db"
    conn = connect(db_path)
    conn.execute("DROP INDEX idx_processed_files_context_sha256")
    conn.execute("CREATE UNIQUE INDEX idx_processed_files_sha256 ON processed_files(file_sha256)")
    conn.execute(
        """
        INSERT INTO processed_files (
            file_id, file_path, file_name, file_sha256, detected_at, status, created_at, updated_at
        ) VALUES ('legacy', 'C:/ventas.csv', 'ventas.csv', 'legacy-sha', '2026-08-30', 'sent', '2026-08-30', '2026-08-30')
        """
    )
    conn.commit()
    conn.close()

    migrated = connect(db_path)
    try:
        assert migrated.execute("SELECT COUNT(*) FROM processed_files WHERE file_id = 'legacy'").fetchone()[0] == 1
        indexes = {row[1] for row in migrated.execute("PRAGMA index_list(processed_files)")}
        assert "idx_processed_files_sha256" not in indexes
        assert "idx_processed_files_context_sha256" in indexes
        assert migrated.execute("PRAGMA quick_check").fetchone()[0] == "ok"
    finally:
        migrated.close()
